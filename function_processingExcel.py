import openpyxl
from openpyxl.styles import Font, PatternFill


def createExcel(FloorDataSheet, ExcelFileName: str):
    wb = openpyxl.Workbook()
    s1 = wb.active 

    s1['A1'].value = '樓層'
    for idx, f in enumerate(FloorDataSheet):
        cell = str(f"A{idx+2}")
        s1[cell].value = FloorDataSheet[idx].name

    s1['B1'].value = '樓地版面積'
    s1['B1'].fill = PatternFill(fill_type="solid", fgColor="FDE9D9")
    for idx, f in enumerate(FloorDataSheet):
        cell = str(f"B{idx+2}")
        s1[cell].value = FloorDataSheet[idx].area_floor
        s1[cell].fill = PatternFill(fill_type="solid", fgColor="FDE9D9")

    s1['C1'].value = '梯廳面積'
    s1["C1"].fill = PatternFill(fill_type="solid", fgColor="DAEEF3")
    for idx, f in enumerate(FloorDataSheet):
        cell = str(f"C{idx+2}")
        s1[cell].value = FloorDataSheet[idx].area_hall
        s1[cell].fill = PatternFill(fill_type="solid", fgColor="DAEEF3")

    s1['D1'].value = '機電面積'
    s1["D1"].fill = PatternFill(fill_type="solid", fgColor="E4DFEC")
    for idx, f in enumerate(FloorDataSheet):
        cell = str(f"D{idx+2}")
        s1[cell].value = FloorDataSheet[idx].area_electromechanical
        s1[cell].fill = PatternFill(fill_type="solid", fgColor="E4DFEC")

    s1['E1'].value = '管委會面積'
    s1['E1'].fill = PatternFill(fill_type="solid", fgColor="EBF1DE")
    for idx, f in enumerate(FloorDataSheet):
        cell = str(f"E{idx+2}")
        s1[cell].value = FloorDataSheet[idx].area_committee
        s1[cell].fill = PatternFill(fill_type="solid", fgColor="EBF1DE")

    s1['F1'].value = '陽台面積'
    s1['F1'].fill = PatternFill(fill_type="solid", fgColor="F2DCDB")
    for idx, f in enumerate(FloorDataSheet):
        cell = str(f"F{idx+2}")
        s1[cell].value = FloorDataSheet[idx].area_balcony
        s1[cell].fill = PatternFill(fill_type="solid", fgColor="F2DCDB")

    s1['G1'].value = '室內停車位與車道面積'
    s1['G1'].fill = PatternFill(fill_type="solid", fgColor="F2F2F2")
    for idx, f in enumerate(FloorDataSheet):
        cell = str(f"G{idx+2}")
        s1[cell].value = FloorDataSheet[idx].area_carramp
        s1[cell].fill = PatternFill(fill_type="solid", fgColor="F2F2F2")

    s1['H1'].value = '梯廳反計容積'
    s1['H1'].fill = PatternFill(fill_type="solid", fgColor="FCD5B4")
    for idx, f in enumerate(FloorDataSheet):
        cell = str(f"H{idx+2}")
        s1[cell].value = FloorDataSheet[idx].over_hall
        s1[cell].fill = PatternFill(fill_type="solid", fgColor="FCD5B4")
    
    s1['I1'].value = '陽台反計容積'
    s1['I1'].fill = PatternFill(fill_type="solid", fgColor="E6B8B7")
    for idx, f in enumerate(FloorDataSheet):
        cell = str(f"I{idx+2}")
        s1[cell].value = FloorDataSheet[idx].over_balcony
        s1[cell].fill = PatternFill(fill_type="solid", fgColor="E6B8B7")

    s1['J1'].value = '梯廳與陽台反計容積'
    s1['J1'].fill = PatternFill(fill_type="solid", fgColor="D8E4BC")
    for idx, f in enumerate(FloorDataSheet):
        cell = str(f"J{idx+2}")
        s1[cell].value = FloorDataSheet[idx].over_hallBalcony
        s1[cell].fill = PatternFill(fill_type="solid", fgColor="D8E4BC")

    s1['K1'].value = '容積樓地板面積'
    for idx, f in enumerate(FloorDataSheet):
        cell = str(f"K{idx+2}")
        s1[cell].value = FloorDataSheet[idx].area_volumn

    ExcelFileName = ExcelFileName + ".xlsx"  
    wb.save(ExcelFileName)

