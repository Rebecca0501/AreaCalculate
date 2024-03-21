import os
import pathlib
import numpy as np
import ezdxf
from ezdxf.enums import TextEntityAlignment
from object import CPoint
from object import CLine
from object import CSqure
from object import CPointVector
import openpyxl
from openpyxl.styles import Font, PatternFill

class floor_imformation():
    def __init__(self, x_min:float, x_max:float, y_min:float, y_max:float):
        self.x_min = x_min
        self.x_max = x_max
        self.y_min = y_min
        self.y_max = y_max
        self.write_line = 0
        self.font_size = abs(y_max-y_min)/50
        self.write_point = CPoint(self.x_max,self.y_max)

        self.area_floor = 0.00
        self.area_floorRemoveCar = 0.00
        self.area_hall = 0.00
        self.area_committee = 0.00
        self.area_electromechanical = 0.00
        self.area_carramp = 0.00
        self.area_balcony = 0.00
        self.area_volumn = 0.00

        self.over_hall =  0.00
        self.over_balcony =  0.00
        self.over_hallBalcony =  0.00

        self.name = "new floor"

    # def __str__(self):
    #     return f"Vector: [{self.vx}, {self.vy}]"
        
class point_list():
    def __init__(self, LWP_point:[]):
        self.LWP_point = LWP_point
    
    def result(self):
        points = []
        for inx, p in enumerate(self.LWP_point):
            points.append(CPoint(p[0],p[1]))
        return points

class detA_sum():
    def __init__(self, point_num:int, point):
        self.point_num = point_num
        self.point = point
    
    def result(self):
        origin_vector = []
        sum_detA = 0.000
        for i in range(self.point_num):
            vector_i = CPointVector(self.point[i],self.point[0])
            origin_vector.append(vector_i)

        for i in range(len(origin_vector)-1):
            matrix = [[origin_vector[i].vx,origin_vector[i].vy],
                      [origin_vector[i+1].vx,origin_vector[i+1].vy]]
            #print(f"Matrix {i+1} = {matrix}")
            s= np.linalg.det(matrix)*0.5
            sum_detA += s
        return sum_detA

def getFloorBorder(msp, LayerName:str, FloorDataSheet):
    for LWPOLYLINE in msp.query("LWPOLYLINE[layer=='{}']".format(LayerName)):
        LWP_point = LWPOLYLINE.get_points()
        p1 = CPoint(LWP_point[0][0],LWP_point[0][1])
        p2 = CPoint(LWP_point[1][0],LWP_point[1][1])
        p3 = CPoint(LWP_point[2][0],LWP_point[2][1])
        p4 = CPoint(LWP_point[3][0],LWP_point[3][1])
        select_area = CSqure(CLine(p1,p2),CLine(p4,p3))
        
        x_min = select_area.x_min()
        x_max = select_area.x_max()
        y_min = select_area.y_min()
        y_max = select_area.y_max()
        new_floor = floor_imformation(x_min,x_max,y_min,y_max)
        FloorDataSheet.append(new_floor)

def getFloorName(msp, LayerName:str, FloorDataSheet):
    for TEXT in msp.query("TEXT[layer=='{}']".format(LayerName)):
        TextCoordinate = TEXT.dxf.insert
        TextCoordiPoint = CPoint(TextCoordinate[0],TextCoordinate[1])
        text_py = TEXT.dxf.text
        for idx, FloorInfo in enumerate(FloorDataSheet):
            if FloorInfo.x_min < TextCoordiPoint.x and TextCoordiPoint.x < FloorInfo.x_max and FloorInfo.y_min < TextCoordiPoint.y and TextCoordiPoint.y < FloorInfo.y_max:
                FloorDataSheet[idx].name = text_py
                break

    for MTEXT in msp.query("MTEXT[layer=='{}']".format(LayerName)):
        TextCoordinate = MTEXT.dxf.insert
        TextCoordiPoint = CPoint(TextCoordinate[0],TextCoordinate[1])
        text_py = MTEXT.dxf.text
        for idx, FloorInfo in enumerate(FloorDataSheet):
            if FloorInfo.x_min < TextCoordiPoint.x and TextCoordiPoint.x < FloorInfo.x_max and FloorInfo.y_min < TextCoordiPoint.y and TextCoordiPoint.y < FloorInfo.y_max:
                FloorDataSheet[idx].name = text_py
                break

def getArea(msp, LayerName:str, FloorDataSheet, attribute_name):
    for LWPOLYLINE in msp.query("LWPOLYLINE[layer=='{}']".format(LayerName)):
        points = point_list(LWPOLYLINE.get_points()).result()
        IntPointNum = len(points)
        for idx, f in enumerate(FloorDataSheet):
            if f.x_min < points[0].x and points[0].x < f.x_max and f.y_min < points[0].y and points[0].y < f.y_max:
                #****************Be careful unit conversion**********
                sum_detA = abs(detA_sum(IntPointNum,points).result())*0.0001
                #****************Be careful unit conversion**********
                setattr(FloorDataSheet[idx], attribute_name, getattr(FloorDataSheet[idx], attribute_name) + sum_detA)
                break

def createExcel(FloorDataSheet, ExcelFileName: str):
    wb = openpyxl.Workbook()
    s1 = wb.active 

    s1['A1'].value = '樓層'
    for idx, f in enumerate(FloorDataSheet):
        cell = str(f"A{idx+2}")
        s1[cell].value = FloorDataSheet[idx].name

    s1['B1'].value = '樓地版面積'
    s1['B1'].fill = PatternFill(fill_type="solid", fgColor="FDE9D9")
    for idx, f in enumerate(FloorDataSheet):
        cell = str(f"B{idx+2}")
        s1[cell].value = FloorDataSheet[idx].area_floor
        s1[cell].fill = PatternFill(fill_type="solid", fgColor="FDE9D9")

    s1['C1'].value = '梯廳面積'
    s1["C1"].fill = PatternFill(fill_type="solid", fgColor="DAEEF3")
    for idx, f in enumerate(FloorDataSheet):
        cell = str(f"C{idx+2}")
        s1[cell].value = FloorDataSheet[idx].area_hall
        s1[cell].fill = PatternFill(fill_type="solid", fgColor="DAEEF3")

    s1['D1'].value = '機電面積'
    s1["D1"].fill = PatternFill(fill_type="solid", fgColor="E4DFEC")
    for idx, f in enumerate(FloorDataSheet):
        cell = str(f"D{idx+2}")
        s1[cell].value = FloorDataSheet[idx].area_electromechanical
        s1[cell].fill = PatternFill(fill_type="solid", fgColor="E4DFEC")

    s1['E1'].value = '管委會面積'
    s1['E1'].fill = PatternFill(fill_type="solid", fgColor="EBF1DE")
    for idx, f in enumerate(FloorDataSheet):
        cell = str(f"E{idx+2}")
        s1[cell].value = FloorDataSheet[idx].area_committee
        s1[cell].fill = PatternFill(fill_type="solid", fgColor="EBF1DE")

    s1['F1'].value = '陽台面積'
    s1['F1'].fill = PatternFill(fill_type="solid", fgColor="F2DCDB")
    for idx, f in enumerate(FloorDataSheet):
        cell = str(f"F{idx+2}")
        s1[cell].value = FloorDataSheet[idx].area_balcony
        s1[cell].fill = PatternFill(fill_type="solid", fgColor="F2DCDB")

    s1['G1'].value = '室內停車位與車道面積'
    s1['G1'].fill = PatternFill(fill_type="solid", fgColor="F2F2F2")
    for idx, f in enumerate(FloorDataSheet):
        cell = str(f"G{idx+2}")
        s1[cell].value = FloorDataSheet[idx].area_carramp
        s1[cell].fill = PatternFill(fill_type="solid", fgColor="F2F2F2")

    s1['H1'].value = '梯廳反計容積'
    s1['H1'].fill = PatternFill(fill_type="solid", fgColor="FCD5B4")
    for idx, f in enumerate(FloorDataSheet):
        cell = str(f"H{idx+2}")
        s1[cell].value = FloorDataSheet[idx].over_hall
        s1[cell].fill = PatternFill(fill_type="solid", fgColor="FCD5B4")
    
    s1['I1'].value = '陽台反計容積'
    s1['I1'].fill = PatternFill(fill_type="solid", fgColor="E6B8B7")
    for idx, f in enumerate(FloorDataSheet):
        cell = str(f"I{idx+2}")
        s1[cell].value = FloorDataSheet[idx].over_balcony
        s1[cell].fill = PatternFill(fill_type="solid", fgColor="E6B8B7")

    s1['J1'].value = '梯廳與陽台反計容積'
    s1['J1'].fill = PatternFill(fill_type="solid", fgColor="D8E4BC")
    for idx, f in enumerate(FloorDataSheet):
        cell = str(f"J{idx+2}")
        s1[cell].value = FloorDataSheet[idx].over_hallBalcony
        s1[cell].fill = PatternFill(fill_type="solid", fgColor="D8E4BC")

    s1['K1'].value = '容積樓地板面積'
    for idx, f in enumerate(FloorDataSheet):
        cell = str(f"K{idx+2}")
        s1[cell].value = FloorDataSheet[idx].area_volumn

    ExcelFileName = ExcelFileName + ".xlsx"  
    wb.save(ExcelFileName)

def addTextLine(msp, FloorInfo, Content: str, line:int):
    msp.add_text(
        Content,
        height=FloorInfo.font_size,
    ).set_placement((FloorInfo.write_point.x, FloorInfo.write_point.y), align=TextEntityAlignment.LEFT)
    FloorInfo.write_point.y -= (FloorInfo.font_size*1.5*line)

def createCalculateFormula(FloorDataSheet):
    for idx, FloorInfo in enumerate(FloorDataSheet):

        FloorInfo.area_floorRemoveCar = FloorInfo.area_floor - FloorInfo.area_carramp

        FloorInfo.over_hall =  max(FloorInfo.area_hall - (FloorInfo.area_floorRemoveCar*0.1),0)

        FloorInfo.over_balcony =  max(FloorInfo.area_balcony - (FloorInfo.area_floorRemoveCar*0.1),0)

        FloorInfo.over_hallBalcony = max(FloorInfo.area_hall + FloorInfo.area_balcony - ((FloorInfo.area_floorRemoveCar)*0.15) - FloorInfo.over_hall - FloorInfo.over_balcony,0)

        FloorInfo.area_volumn = FloorInfo.area_floor-FloorInfo.area_hall-FloorInfo.area_committee-FloorInfo.area_electromechanical-FloorInfo.area_carramp + FloorInfo.over_hall + FloorInfo.over_balcony + FloorInfo.over_hallBalcony

        if FloorInfo.area_floor >0:
            content = str(f"樓地板面積: {round(FloorInfo.area_floor,2)}")
            addTextLine(msp, FloorInfo, content, 2)

        if FloorInfo.area_carramp >0.:
            content = str(f"室內停車&車道面積: {round(FloorInfo.area_carramp,2)}")
            addTextLine(msp, FloorInfo, content, 2)

            content = str("扣除室內停車面積之樓地板面積:")
            addTextLine(msp, FloorInfo, content, 1)
            
            content = str(f"= {round(FloorInfo.area_floor,2)}(樓地板面積) - {round(FloorInfo.area_carramp,2)}(車道面積) = {round(FloorInfo.area_floorRemoveCar,2)}")
            addTextLine(msp, FloorInfo, content, 2)

        if FloorInfo.area_hall >0.:
            content = str(f"梯廳面積: {round(FloorInfo.area_hall,2)}")
            addTextLine(msp, FloorInfo, content, 1)

            if FloorInfo.over_hall<0.001:
                content = str(f"{round(FloorInfo.area_hall,2)} < {round(FloorInfo.area_floorRemoveCar,2)}*0.1 = {round(FloorInfo.area_floorRemoveCar*0.1,2)}...OK")
                addTextLine(msp, FloorInfo, content, 2)
            else:
                content = str(f"{round(FloorInfo.area_hall,2)} > {round(FloorInfo.area_floorRemoveCar,2)}*0.1 = {round(FloorInfo.area_floorRemoveCar*0.1,2)}")
                addTextLine(msp, FloorInfo, content, 1)

                content = str(f"{round(FloorInfo.area_hall,2)} - {round(FloorInfo.area_floorRemoveCar*0.1,2)} = {round(FloorInfo.over_hall,2)}...計入容積")
                addTextLine(msp, FloorInfo, content, 2)

        if FloorInfo.area_committee >0.:
            content = str(f"管委會面積: {round(FloorInfo.area_committee,2)}")
            addTextLine(msp, FloorInfo, content, 2)

        if FloorInfo.area_electromechanical >0.:
            content = str(f"機電面積: {round(FloorInfo.area_electromechanical,2)}")
            addTextLine(msp, FloorInfo, content, 2)

        if FloorInfo.area_balcony >0.:
            content = str(f"陽台面積: {round(FloorInfo.area_balcony,2)}")
            addTextLine(msp, FloorInfo, content, 1)
            
            if FloorInfo.over_balcony<0.001:
                content = str(f"{round(FloorInfo.area_balcony,2)} < {round(FloorInfo.area_floorRemoveCar,2)}*0.1 = {round(FloorInfo.area_floorRemoveCar*0.1,2)}...OK")
                addTextLine(msp, FloorInfo, content, 2)
                
            else:
                content = str(f"{round(FloorInfo.area_balcony,2)} > {round(FloorInfo.area_floorRemoveCar,2)}*0.1 = {round(FloorInfo.area_floorRemoveCar*0.1,2)}")
                addTextLine(msp, FloorInfo, content, 1)
                
                content = str(f"{round(FloorInfo.area_balcony,2)} - {round(FloorInfo.area_floorRemoveCar*0.1,2)} = {round(FloorInfo.over_balcony,2)}...計入容積")
                addTextLine(msp, FloorInfo, content, 2)

        content = str("梯廳+陽台面積檢討")
        addTextLine(msp, FloorInfo, content, 1)

        content = str(f"{round(FloorInfo.area_hall,2)}+{round(FloorInfo.area_balcony,2)}={round(FloorInfo.area_hall,2)+round(FloorInfo.area_balcony,2)}")
        addTextLine(msp, FloorInfo, content, 1)
        
        if (FloorInfo.area_hall + FloorInfo.area_balcony) < (FloorInfo.area_floorRemoveCar)*0.15:
            content = str(f"{round(FloorInfo.area_hall,2)+round(FloorInfo.area_balcony,2)} < {round(FloorInfo.area_floorRemoveCar,2)}*0.15= {round(FloorInfo.area_floorRemoveCar*0.15,2)} ...OK")
            addTextLine(msp, FloorInfo, content, 2)
        else:
            content = str(f"{round(FloorInfo.area_hall,2)+round(FloorInfo.area_balcony,2)} > {round(FloorInfo.area_floorRemoveCar,2)}*0.15= {round(FloorInfo.area_floorRemoveCar*0.15,2)}")
            addTextLine(msp, FloorInfo, content, 1)

            content = str(f"{round(FloorInfo.area_hall,2)+round(FloorInfo.area_balcony,2)} - {round(FloorInfo.area_floorRemoveCar*0.15,2)} - {round(FloorInfo.over_hall,2)}(10%梯廳超出) - {round(FloorInfo.over_balcony,2)}(10%陽台超出) = {round(FloorInfo.over_hallBalcony,2)}...計入容積")
            addTextLine(msp, FloorInfo, content, 2)

        content = str("容積樓地板面積")
        addTextLine(msp, FloorInfo, content, 1)

        content = str(f"= {round(FloorInfo.area_floor,2)}(樓地板面積) - {round(FloorInfo.area_hall,2)}(梯廳面積) - {round(FloorInfo.area_committee,2)}(管委會面積) - {round(FloorInfo.area_electromechanical,2)}(機電面積) - {round(FloorInfo.area_carramp,2)}(車道面積) + {round(FloorInfo.over_hall,2)}(10%梯廳超出) + {round(FloorInfo.over_balcony,2)}(10%陽台超出) + {round(FloorInfo.over_hallBalcony,2)}(15%梯廳陽台超出)")
        addTextLine(msp, FloorInfo, content, 1)

        content = str(f"= {round(FloorInfo.area_volumn,2)}")
        addTextLine(msp, FloorInfo, content, 1)

if __name__=="__main__":

    #Step1. Get CAD file path
    str_CAD_file_name="example.dxf"
    pathInputFilePath=pathlib.Path(os.getcwd())
    pathInputFilePath=pathInputFilePath.joinpath(str_CAD_file_name)
    input_file_path = str(pathInputFilePath)

    #Step2. Open CAD file and model space
    doc = ezdxf.readfile(input_file_path)
    msp = doc.modelspace()

    #Step3. create floor data sheet record imformation of select area
    FloorDataSheet = []
    getFloorBorder(msp, "SelectArea", FloorDataSheet)
    getFloorName(msp, "floor_name", FloorDataSheet)
    getArea(msp, "41-樓地板面積", FloorDataSheet, "area_floor")
    getArea(msp, "42-梯廳面積", FloorDataSheet, "area_hall")
    getArea(msp, "43-管委會面積", FloorDataSheet, "area_committee")
    getArea(msp, "44-機電面積", FloorDataSheet, "area_electromechanical")
    getArea(msp, "45-車道面積", FloorDataSheet, "area_carramp")
    getArea(msp, "46-陽台面積", FloorDataSheet, "area_balcony")

    #Step4. Add floor information on the new dxf file. 
    createCalculateFormula(FloorDataSheet)

    #Step5. Save processed file
    file_name_parts = input_file_path.split('.')
    output_file_path0 = file_name_parts[0] + "-after.dxf"
    doc.saveas(output_file_path0)

    #Step6. Save value in excel file
    createExcel(FloorDataSheet, "面積計算表")
